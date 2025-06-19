import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

st.set_page_config(layout="wide")
st.title("メニューランキングアンケート")

# メニュー詳細を分けて管理
menus = [
    {"番号": 1, "バーガー": "チーズバーガー", "サイド": "フライドポテト", "ドリンク": "ホットコーヒー", "値段": 570},
    {"番号": 2, "バーガー": "チーズバーガー", "サイド": "コールスローサラダ", "ドリンク": "コーラ", "値段": 680},
    {"番号": 3, "バーガー": "チーズバーガー", "サイド": "グリーンサラダ", "ドリンク": "オレンジジュース", "値段": 620},
    {"番号": 4, "バーガー": "フィッシュバーガー", "サイド": "フライドポテト", "ドリンク": "コーラ", "値段": 620},
    {"番号": 5, "バーガー": "フィッシュバーガー", "サイド": "コールスローサラダ", "ドリンク": "オレンジジュース", "値段": 570},
    {"番号": 6, "バーガー": "フィッシュバーガー", "サイド": "グリーンサラダ", "ドリンク": "ホットコーヒー", "値段": 680},
    {"番号": 7, "バーガー": "ホットサンド", "サイド": "フライドポテト", "ドリンク": "オレンジジュース", "値段": 680},
    {"番号": 8, "バーガー": "ホットサンド", "サイド": "コールスローサラダ", "ドリンク": "ホットコーヒー", "値段": 620},
    {"番号": 9, "バーガー": "ホットサンド", "サイド": "グリーンサラダ", "ドリンク": "コーラ", "値段": 570},
]

st.write("あなたのお名前を入力してください。")
name = st.text_input("名前")

st.write("各メニューの順位を1〜9で選んでください（順位は重複できません）")

# 順位選択肢
rank_options = [str(i) for i in range(1, 10)]

# 3列×3行でカード表示
cols = st.columns(3)

# 順位選択用の辞書初期化
rank_selected = {}

for idx, menu in enumerate(menus):
    col = cols[idx % 3]
    with col:
        st.markdown(
            f"""
            <div style="border:2px solid #4CAF50; border-radius:10px; padding:15px; margin-bottom:10px;">
                <h3>[{menu['番号']}] メニュー</h3>
                <p><b>バーガー：</b>{menu['バーガー']}</p>
                <p><b>サイド：</b>{menu['サイド']}</p>
                <p><b>ドリンク：</b>{menu['ドリンク']}</p>
                <p><b>値段：</b>{menu['値段']}円</p>
            </div>
            """, unsafe_allow_html=True
        )
        rank_selected[menu['番号']] = col.selectbox(
            f"順位を選択 [{menu['番号']}]",
            options=rank_options,
            key=f"rank_{menu['番号']}"
        )

# 送信ボタン
if st.button("送信"):

    # 入力チェック
    if not name.strip():
        st.error("名前を入力してください。")
    else:
        # 重複順位チェック
        ranks = list(rank_selected.values())
        if len(set(ranks)) != 9:
            st.error("順位は重複できません。すべて違う順位を選んでください。")
        else:
            # Excelファイル書き込み
            filepath = "results.xlsx"
            if os.path.exists(filepath):
                wb = load_workbook(filepath)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # 1行目：メニュー番号、2～10行：[1]～[9]
                ws["A1"] = "メニュー番号"
                for i in range(9):
                    ws.cell(row=i + 2, column=1, value=f"[{i+1}]")
                # 1行目に名前（B1～AE1は30人分）
                for i in range(30):
                    col_letter = get_column_letter(i + 2)  # B列から
                    ws[f"{col_letter}1"] = f"{i+1}人目"
                # タイトルと計算式
                ws["AF1"] = "合計"
                ws["AG1"] = "基点"
                ws["AH1"] = "得点"
                ws["AI1"] = "順位"
                # 計算式を2行目に設定
                ws["AF2"] = "=SUM(B2:AE2)"
                ws["AG2"] = "=COUNT(B2:AE2)*10"
                ws["AH2"] = "=AG2-AF2"

            # 空いている列を探す（B～AE列、30人分）
            empty_col = None
            for col_idx in range(2, 32):  # B(2)～AE(31)
                if ws.cell(row=1, column=col_idx).value is None:
                    empty_col = col_idx
                    break

            if empty_col is None:
                st.error("30人分のアンケートが埋まっています。")
            else:
                # 名前を1行目に入れる
                ws.cell(row=1, column=empty_col, value=name)
                # 順位を2～10行に書き込む
                for i, menu in enumerate(menus):
                    rank_value = int(rank_selected[menu["番号"]])
                    ws.cell(row=i + 2, column=empty_col, value=rank_value)

                # 計算式のコピー(3行目以降)
                for row in range(3, 11):
                    ws[f"AF{row}"] = ws["AF2"].value
                    ws[f"AG{row}"] = ws["AG2"].value
                    ws[f"AH{row}"] = ws["AH2"].value

                wb.save(filepath)
                st.success(f"送信ありがとうございました！あなたは{empty_col - 1}人目の回答者です。")

